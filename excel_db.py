import openpyxl
import datetime
import os


class ExcelWorkBook:
    file_path = ""

    def __init__(self) -> None:
        try:
            self.wb = openpyxl.load_workbook(r"./attendance/data.xlsx")
        except:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet["A1"] = "id"
            sheet["B1"] = "face_id"
            sheet["C1"] = "name"
            sheet["D1"] = "branch"
            sheet["E1"] = "class"
            # sheet["F1"] = "created_at"
            # sheet["F1"] = "updated_at"
            wb.save("./attendance/data.xlsx")
            self.wb = openpyxl.load_workbook(r"./attendance/data.xlsx")

    def add_user(self, id, face_id, name, branch, class_name):
        sheet = self.wb.active
        cur_row = self.current_row()
        sheet.cell(row=cur_row + 1, column=1).value = id
        sheet.cell(row=cur_row + 1, column=2).value = face_id
        sheet.cell(row=cur_row + 1, column=3).value = name
        sheet.cell(row=cur_row + 1, column=4).value = branch
        sheet.cell(row=cur_row + 1, column=5).value = class_name
        # sheet.cell(row=cur_row + 1, column=6).value = datetime.datetime.now()
        self.wb.save("./attendance/data.xlsx")

    def current_row(self):
        return self.wb.active.max_row

    def get_user(self):
        sheet = self.wb.active
        list = []
        for row in sheet["C2:C" + str(self.current_row())]:
            list.append([x.value for x in row][0])
        return list

    def makeAttendace(self, id):
        today = datetime.date.today()
        filepath = "attendance/" + str(today.month) + str(today.year) + ".xlsx"
        # check if the file exists
        if not os.path.isfile(filepath):
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet["A1"] = "roll no"
            wb.save(filepath)

        wb = openpyxl.load_workbook(
            r"attendance/" + str(today.month) + str(today.year) + ".xlsx"
        )
        """_summary_
        """
        sheet = wb.active
        cur_row = int(id) + 1
        # print(cur_row)
        if sheet.cell(row=1, column=sheet.max_column).value != str(today):
            sheet.cell(row=1, column=sheet.max_column + 1).value = str(today)
        sheet.cell(row=cur_row + 1, column=sheet.max_column - 1).value = int(id) + 1
        # print(cur_row + 1, " ", int(sheet.max_column))
        sheet.cell(row=cur_row + 1, column=sheet.max_column).value = "P"
        wb.save(filepath)


# debug and testing
# ExcelWorkBook().makeAttendace(10)
